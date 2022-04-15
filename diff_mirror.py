
import os
import collections
import rpmfile
import json
import argparse
import logging
from openpyxl  import Workbook

def get_rpm_info(rf):
    r=collections.namedtuple('rpm', ['name', 'arch','version'])
    try:
        with rpmfile.open(rf) as rpm:
            r.arch = rpm.headers.get('arch').decode("utf-8")
            r.name = rpm.headers.get('name').decode("utf-8")
            r.version = rpm.headers.get('version').decode("utf-8")
    except:
        return  None

    return r

def get_rpm_list(rpm_dir_path):
    rpm_list = []
    rpm_file_paths = get_rpm_file_paths(rpm_dir_path)
    for rpm_file in rpm_file_paths:
        r = get_rpm_info(rpm_file)
        if r == None:
            continue
        ele="{}/{}".format(r.name,r.version)
        if ele not in rpm_list:
            rpm_list.append(ele)
    return rpm_list

def diff_rpm_list(rpm_list1,rpm_list2):
    data={
        "version_not_match": {},
        "rpm_list1_only_exist": [],
        "rpm_list2_only_exist": [],
        "match":[]
    }
    for r1 in rpm_list1:
        for i,r2 in enumerate(rpm_list2):
            # 名字与版本都不同
            if r1!=r2:
                # 名字相同 版本不同
                if os.path.dirname(r1)==os.path.dirname(r2):
                    data['version_not_match'][os.path.dirname(r1)]=[r1,r2]
                    break

                # 遍历到列表最后一个元素，仍未匹配名字成功，判断包在r2列表不存在
                else:
                    if i == len(rpm_list2) - 1:
                        data['rpm_list1_only_exist'].append(r1)
            else:
                data["match"].append(r1)
                break


    for r1 in rpm_list2:
        for i,r2 in enumerate(rpm_list1):
            if os.path.dirname(r1) == os.path.dirname(r2):
                break
                # 遍历到列表最后一个元素，仍未匹配名字成功，判断包在r2列表不存在
            else:
                if i == len(rpm_list1) - 1 :
                    data['rpm_list2_only_exist'].append(r1)

    # 去重
    data["match"]=list(set(data["match"]))

    return data

def get_rpm_file_paths(file_dir):
    for root,_,files in os.walk(file_dir):
        for f in files:
            if os.path.splitext(f)[-1] == '.rpm':
                yield "{}/{}".format(root,f)

def write_to_xlsx(data,dest_filename):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "version_not_match"
    _ = ws1.cell(column=1, row=1, value="package name")
    _ = ws1.cell(column=2, row=1, value="version1")
    _ = ws1.cell(column=3, row=1, value="version2")
    row = 2
    for name,vs in data['version_not_match'].items():
        _ = ws1.cell(column=1,row=row,value=name)
        _ = ws1.cell(column=2, row=row, value=os.path.basename(vs[0]))
        _ = ws1.cell(column=3, row=row, value=os.path.basename(vs[1]))
        row+=1

    ws2 = wb.create_sheet(title='match')
    _ = ws2.cell(column=1,row=1,value="package name")
    _ = ws2.cell(column=2,row=2,value="version")

    row = 2
    for rv in data['match']:
        _ = ws2.cell(column=1, row=row, value=os.path.dirname(rv))
        _ = ws2.cell(column=2, row=row, value=os.path.basename(rv))
        row += 1

    ws3 = wb.create_sheet(title='rpm_list1_only_exist')
    _ = ws3.cell(column=1, row=1, value="package name")
    _ = ws3.cell(column=2, row=2, value="version")
    row = 2
    for rv in data['rpm_list1_only_exist']:
        _ = ws3.cell(column=1, row=row, value=os.path.dirname(rv))
        _ = ws3.cell(column=2, row=row, value=os.path.basename(rv))
        row += 1

    ws4 = wb.create_sheet(title='rpm_list2_only_exist')
    _ = ws4.cell(column=1, row=1, value="package name")
    _ = ws4.cell(column=2, row=2, value="version")
    row = 2
    for rv in data['rpm_list2_only_exist']:
        _ = ws4.cell(column=1, row=row, value=os.path.dirname(rv))
        _ = ws4.cell(column=2, row=row, value=os.path.basename(rv))
        row += 1
    wb.save(filename=dest_filename)

def main():
    # 日志相关配置
    level = logging.INFO
    datefmt = '%m/%d/%Y %I:%M:%S %p'
    format = '%(asctime)s:%(levelname)s:%(message)s'
    logging.basicConfig(format=format, datefmt=datefmt, filemode='a', level=level)

    parser = argparse.ArgumentParser(description='diff mirror')
    parser.add_argument('--rpm_list_path1', type=str,help='rpm_list_path1')
    parser.add_argument('--rpm_list_path2', type=str,help='rpm_list_path2')
    args = parser.parse_args()

    rpm_list_path1=args.rpm_list_path1
    rpm_list_path2 = args.rpm_list_path2
    logging.info("rpm_list_path1={}".format(rpm_list_path1))
    logging.info("rpm_list_path2={}".format(rpm_list_path2))

    rpm_list_path1_rpm=get_rpm_list(rpm_list_path1)
    rpm_list_path2_rpm =get_rpm_list(rpm_list_path2)

    logging.info("rpm_list_path1: {} rpm file".format(len(rpm_list_path1_rpm)))
    logging.info("rpm_list_path2: {} rpm file".format(len(rpm_list_path2_rpm)))
    if len(rpm_list_path2_rpm)==0 or len(rpm_list_path1_rpm)==0:
        exit(1)

    logging.info("start diff rpm list")
    d=diff_rpm_list(rpm_list_path1_rpm,rpm_list_path2_rpm)

    logging.info("write to xlsx")
    write_to_xlsx(d,"diff.xlsx")
    print(json.dumps(d,indent=4,ensure_ascii=False))

if __name__ == '__main__':
    main()
