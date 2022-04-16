
import os
import collections
import rpmfile
import json
import argparse
import logging
from openpyxl  import Workbook

def get_rpm_info(rf):
    r=collections.namedtuple('rpm', ['name', 'arch','version'])
    try:
        with rpmfile.open(rf) as rpm:
            r.arch = rpm.headers.get('arch').decode("utf-8")
            r.name = rpm.headers.get('name').decode("utf-8")
            r.version = rpm.headers.get('version').decode("utf-8")
    except:
        return  None

    return r

def get_rpm_list(rpm_dir_path):
    rpm_list = []
    rpm_file_paths = get_rpm_file_paths(rpm_dir_path)
    for rpm_file in rpm_file_paths:
        r = get_rpm_info(rpm_file)
        if r == None:
            continue
        ele="{}/{}".format(r.name,r.version)
        if ele not in rpm_list:
            rpm_list.append(ele)
    return rpm_list


def diff_rpm_list(rpm_lists):
    data={
    }
    package_names=[]
    for rpm_list in rpm_lists:
        for r in rpm_list:
            package_name=os.path.dirname(r)
            if package_name not in package_names:
                package_names.append(package_name)


    for package_name in package_names:
        data[package_name] = []
        for i,rpm_list in enumerate(rpm_lists):
            for j,rpm in enumerate(rpm_list):
                if os.path.dirname(rpm) == package_name:
                    version=os.path.basename(rpm)
                    data[package_name].append(version)
                    break
                else:
                    if j== len(rpm_list)-1:
                        data[package_name].append("not exist")

    return data

def get_rpm_file_paths(file_dir):
    for root,_,files in os.walk(file_dir):
        for f in files:
            if os.path.splitext(f)[-1] == '.rpm':
                yield "{}/{}".format(root,f)

def write_to_xlsx(data,dest_filename):
    wb = Workbook()
    ws1 = wb.active

    row=1
    for package_name in data:
        col = 2
        _ = ws1.cell(column=1, row=row, value=package_name)
        for version in data[package_name]:
            _ = ws1.cell(column=col, row=row, value=version)
            col+=1
        row+=1

    wb.save(filename=dest_filename)

def main():
    # 日志相关配置
    level = logging.INFO
    datefmt = '%m/%d/%Y %I:%M:%S %p'
    format = '%(asctime)s:%(levelname)s:%(message)s'
    logging.basicConfig(format=format, datefmt=datefmt, filemode='a', level=level)

    parser = argparse.ArgumentParser(description='diff mirrors')
    parser.add_argument('--rpm_list_paths', type=list,nargs='+',help='rpm_list_paths')
    parser.add_argument('-o', type=str, help='dst_xlsx',default="diff.xlsx")
    args = parser.parse_args()

    dst_filename=args.o

    rpm_lists=[]
    for rl in args.rpm_list_paths:
        rpm_path="".join(rl)
        logging.info("rpm_list_path={}".format(rpm_path))
        rpm_lists.append(get_rpm_list(rpm_path))

    logging.info("start diff rpm list")

    d=diff_rpm_list(rpm_lists)

    logging.info("write to xlsx")
    write_to_xlsx(d,dst_filename)
    print(json.dumps(d,indent=4,ensure_ascii=False))

if __name__ == '__main__':
    main()
